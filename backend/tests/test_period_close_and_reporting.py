import io

from openpyxl import load_workbook


def _gl_account(client, company_id, code, category, forecast_role=None):
    payload = {"code": code, "name": code, "category": category}
    if forecast_role:
        payload["forecast_role"] = forecast_role
    r = client.post(f"/gl-accounts?company_id={company_id}", json=payload)
    assert r.status_code == 200, r.text
    return r.json()["id"]


def _create_entry(client, company_id, lines, entry_date="2026-01-01", reference="Test entry"):
    return client.post(
        f"/journal-entries?company_id={company_id}",
        json={"entry_date": entry_date, "reference": reference, "currency": "USD", "lines": lines},
    )


def _post(client, company_id, lines, entry_date):
    entry = _create_entry(client, company_id, lines, entry_date=entry_date).json()
    r = client.post(f"/journal-entries/{entry['id']}/post")
    assert r.status_code == 200, r.text
    return r.json()


# --- Accruals ------------------------------------------------------------


def test_accrual_can_be_created_and_reversed(client, company):
    company_id = company["id"]
    expense_id = _gl_account(client, company_id, "6300", "expense")
    accrued_liability_id = _gl_account(client, company_id, "2300", "liability")

    r = client.post(
        f"/accruals?company_id={company_id}",
        json={
            "entry_date": "2026-01-31",
            "debit_gl_account_id": expense_id,
            "credit_gl_account_id": accrued_liability_id,
            "amount": 500,
            "reversal_date": "2026-02-01",
            "reference": "Accrued utilities",
        },
    )
    assert r.status_code == 200, r.text
    accrual = r.json()
    assert accrual["reversed"] is False
    assert accrual["amount"] == 500

    actuals = client.get(f"/actuals?company_id={company_id}").json()
    expense_actual = next(a for a in actuals if a["gl_account_id"] == expense_id)
    assert expense_actual["amount"] == 500

    r2 = client.post(f"/accruals/{accrual['id']}/reverse?company_id={company_id}")
    assert r2.status_code == 200, r2.text
    reversed_accrual = r2.json()
    assert reversed_accrual["reversed"] is True
    assert reversed_accrual["reversal_journal_entry_id"] is not None


def test_accrual_amount_must_be_positive(client, company):
    company_id = company["id"]
    expense_id = _gl_account(client, company_id, "6301", "expense")
    liability_id = _gl_account(client, company_id, "2301", "liability")
    r = client.post(
        f"/accruals?company_id={company_id}",
        json={"entry_date": "2026-01-31", "debit_gl_account_id": expense_id, "credit_gl_account_id": liability_id, "amount": -10, "reversal_date": "2026-02-01"},
    )
    assert r.status_code == 422


def test_accrual_reversal_date_must_be_after_entry_date(client, company):
    company_id = company["id"]
    expense_id = _gl_account(client, company_id, "6302", "expense")
    liability_id = _gl_account(client, company_id, "2302", "liability")
    r = client.post(
        f"/accruals?company_id={company_id}",
        json={"entry_date": "2026-01-31", "debit_gl_account_id": expense_id, "credit_gl_account_id": liability_id, "amount": 100, "reversal_date": "2026-01-31"},
    )
    assert r.status_code == 422


def test_cannot_reverse_an_accrual_twice(client, company):
    company_id = company["id"]
    expense_id = _gl_account(client, company_id, "6303", "expense")
    liability_id = _gl_account(client, company_id, "2303", "liability")
    accrual = client.post(
        f"/accruals?company_id={company_id}",
        json={"entry_date": "2026-01-31", "debit_gl_account_id": expense_id, "credit_gl_account_id": liability_id, "amount": 100, "reversal_date": "2026-02-01"},
    ).json()
    client.post(f"/accruals/{accrual['id']}/reverse?company_id={company_id}")
    r = client.post(f"/accruals/{accrual['id']}/reverse?company_id={company_id}")
    assert r.status_code == 409


def test_accrual_due_for_reversal_flag(client, company):
    company_id = company["id"]
    expense_id = _gl_account(client, company_id, "6304", "expense")
    liability_id = _gl_account(client, company_id, "2304", "liability")

    overdue = client.post(
        f"/accruals?company_id={company_id}",
        json={"entry_date": "2019-01-01", "debit_gl_account_id": expense_id, "credit_gl_account_id": liability_id, "amount": 100, "reversal_date": "2020-01-01"},
    ).json()
    not_due = client.post(
        f"/accruals?company_id={company_id}",
        json={"entry_date": "2026-01-01", "debit_gl_account_id": expense_id, "credit_gl_account_id": liability_id, "amount": 100, "reversal_date": "2099-01-01"},
    ).json()

    assert overdue["due_for_reversal"] is True
    assert not_due["due_for_reversal"] is False


def test_accrual_mutations_are_audited(client, company):
    company_id = company["id"]
    expense_id = _gl_account(client, company_id, "6305", "expense")
    liability_id = _gl_account(client, company_id, "2305", "liability")
    accrual = client.post(
        f"/accruals?company_id={company_id}",
        json={"entry_date": "2026-01-31", "debit_gl_account_id": expense_id, "credit_gl_account_id": liability_id, "amount": 100, "reversal_date": "2026-02-01"},
    ).json()
    client.post(f"/accruals/{accrual['id']}/reverse?company_id={company_id}")

    entries = client.get(f"/audit-log?company_id={company_id}&entity_type=accrual").json()
    actions = [e["action"] for e in entries]
    assert "create" in actions
    assert "reverse" in actions


# --- Cash Flow Statement ---------------------------------------------------


def test_cash_flow_statement_proves_itself(client, company):
    company_id = company["id"]
    cash_id = _gl_account(client, company_id, "1400", "asset", forecast_role="cash")
    ar_id = _gl_account(client, company_id, "1401", "asset", forecast_role="accounts_receivable")
    ap_id = _gl_account(client, company_id, "2400", "liability", forecast_role="accounts_payable")
    revenue_id = _gl_account(client, company_id, "4400", "revenue")
    expense_id = _gl_account(client, company_id, "6400", "expense")
    apc_id = _gl_account(client, company_id, "1410", "asset")
    depexp_id = _gl_account(client, company_id, "6410", "expense")
    accumdep_id = _gl_account(client, company_id, "1411", "asset")
    gain_id = _gl_account(client, company_id, "4410", "revenue")
    loss_id = _gl_account(client, company_id, "6411", "expense")

    # Revenue earned on credit: Dr AR 1000 / Cr Revenue 1000
    _post(client, company_id, [{"gl_account_id": ar_id, "debit_amount": 1000, "credit_amount": 0}, {"gl_account_id": revenue_id, "debit_amount": 0, "credit_amount": 1000}], "2026-01-15")
    # Expense incurred, unpaid: Dr Expense 300 / Cr AP 300
    _post(client, company_id, [{"gl_account_id": expense_id, "debit_amount": 300, "credit_amount": 0}, {"gl_account_id": ap_id, "debit_amount": 0, "credit_amount": 300}], "2026-01-20")

    # Asset acquisition funded by cash: 6000, 5yr straight-line
    asset_class = client.post(
        f"/asset-classes?company_id={company_id}",
        json={
            "name": "CF Equipment",
            "apc_gl_account_id": apc_id,
            "depreciation_expense_gl_account_id": depexp_id,
            "accumulated_depreciation_gl_account_id": accumdep_id,
            "disposal_gain_gl_account_id": gain_id,
            "disposal_loss_gl_account_id": loss_id,
            "default_depreciation_method": "straight_line",
            "default_useful_life_years": 5,
        },
    ).json()
    client.post(
        f"/assets?company_id={company_id}",
        json={
            "asset_class_id": asset_class["id"],
            "code": "CF-1",
            "name": "CF Asset",
            "acquisition_date": "2026-01-01",
            "capitalized_cost": 6000,
            "funding_gl_account_id": cash_id,
        },
    )
    client.post(f"/assets/depreciation-run?company_id={company_id}&period=2026-01-01")

    r = client.get(f"/reports/cash-flow-statement?company_id={company_id}&start=2026-01-01&end=2026-01-31")
    assert r.status_code == 200, r.text
    body = r.json()

    assert body["net_income"] == 600.0  # 1000 revenue - 300 expense - 100 depreciation
    assert body["depreciation_add_back"] == 100.0
    assert body["increase_in_receivables"] == 1000.0
    assert body["increase_in_payables"] == 300.0
    assert body["net_operating_cash_flow"] == 0.0  # 600 + 100 - 1000 + 300
    assert body["asset_acquisitions"] == 6000.0
    assert body["disposal_proceeds"] == 0.0
    assert body["net_investing_cash_flow"] == -6000.0
    assert body["net_financing_cash_flow"] == 0.0
    assert body["net_change_in_cash"] == -6000.0
    assert body["opening_cash_balance"] == 0.0
    assert body["closing_cash_balance"] == -6000.0
    assert body["is_proven"] is True


# --- Period close cockpit --------------------------------------------------


def test_period_close_status_flags_whats_missing(client, company):
    company_id = company["id"]
    apc_id = _gl_account(client, company_id, "1420", "asset")
    depexp_id = _gl_account(client, company_id, "6420", "expense")
    accumdep_id = _gl_account(client, company_id, "1421", "asset")
    gain_id = _gl_account(client, company_id, "4420", "revenue")
    loss_id = _gl_account(client, company_id, "6421", "expense")
    cash_id = _gl_account(client, company_id, "1422", "asset")
    revenue_id = _gl_account(client, company_id, "4421", "revenue")
    expense_id = _gl_account(client, company_id, "6422", "expense")

    asset_class = client.post(
        f"/asset-classes?company_id={company_id}",
        json={
            "name": "PC Equipment",
            "apc_gl_account_id": apc_id,
            "depreciation_expense_gl_account_id": depexp_id,
            "accumulated_depreciation_gl_account_id": accumdep_id,
            "disposal_gain_gl_account_id": gain_id,
            "disposal_loss_gl_account_id": loss_id,
        },
    ).json()
    asset = client.post(
        f"/assets?company_id={company_id}",
        json={"asset_class_id": asset_class["id"], "code": "PC-1", "name": "PC Asset", "acquisition_date": "2026-01-01", "capitalized_cost": 1200, "funding_gl_account_id": cash_id},
    ).json()

    # A draft (unposted) entry sitting in the period
    _create_entry(client, company_id, [{"gl_account_id": revenue_id, "debit_amount": 0, "credit_amount": 100}, {"gl_account_id": expense_id, "debit_amount": 100, "credit_amount": 0}], "2026-01-10")

    status = client.get(f"/period-close/status?company_id={company_id}&period=2026-01-01").json()
    assert status["draft_entries_count"] >= 1
    assert status["depreciation_run_done"] is False
    assert any(g["asset_id"] == asset["id"] for g in status["assets_missing_depreciation"])
    assert status["ready_to_close"] is False


def test_period_close_ready_once_everything_done(client, company):
    company_id = company["id"]
    apc_id = _gl_account(client, company_id, "1430", "asset")
    depexp_id = _gl_account(client, company_id, "6430", "expense")
    accumdep_id = _gl_account(client, company_id, "1431", "asset")
    gain_id = _gl_account(client, company_id, "4430", "revenue")
    loss_id = _gl_account(client, company_id, "6431", "expense")
    cash_id = _gl_account(client, company_id, "1432", "asset")

    asset_class = client.post(
        f"/asset-classes?company_id={company_id}",
        json={
            "name": "Ready Equipment",
            "apc_gl_account_id": apc_id,
            "depreciation_expense_gl_account_id": depexp_id,
            "accumulated_depreciation_gl_account_id": accumdep_id,
            "disposal_gain_gl_account_id": gain_id,
            "disposal_loss_gl_account_id": loss_id,
        },
    ).json()
    client.post(
        f"/assets?company_id={company_id}",
        json={"asset_class_id": asset_class["id"], "code": "READY-1", "name": "Ready Asset", "acquisition_date": "2026-01-01", "capitalized_cost": 1200, "funding_gl_account_id": cash_id},
    )
    client.post(f"/assets/depreciation-run?company_id={company_id}&period=2026-01-01")

    status = client.get(f"/period-close/status?company_id={company_id}&period=2026-01-01").json()
    assert status["draft_entries_count"] == 0
    assert status["depreciation_run_done"] is True
    assert status["assets_missing_depreciation"] == []
    assert status["accruals_due_for_reversal"] == 0
    assert status["ready_to_close"] is True


# --- Year-end close ---------------------------------------------------------


def test_year_end_close_zeroes_revenue_and_expense_into_retained_earnings(client, company):
    company_id = company["id"]
    revenue_id = _gl_account(client, company_id, "4500", "revenue")
    expense_id = _gl_account(client, company_id, "6500", "expense")
    cash_id = _gl_account(client, company_id, "1500", "asset")
    retained_earnings_id = _gl_account(client, company_id, "3500", "equity")

    _post(client, company_id, [{"gl_account_id": cash_id, "debit_amount": 1000, "credit_amount": 0}, {"gl_account_id": revenue_id, "debit_amount": 0, "credit_amount": 1000}], "2026-06-01")
    _post(client, company_id, [{"gl_account_id": expense_id, "debit_amount": 400, "credit_amount": 0}, {"gl_account_id": cash_id, "debit_amount": 0, "credit_amount": 400}], "2026-06-15")

    r = client.post(
        f"/year-end/close?company_id={company_id}",
        json={"start": "2026-01-01", "end": "2026-12-31", "retained_earnings_gl_account_id": retained_earnings_id},
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["net_income"] == 600.0

    tb = client.get(f"/trial-balance?company_id={company_id}&as_of=2026-12-31").json()
    revenue_row = next((row for row in tb["rows"] if row["gl_account_id"] == revenue_id), None)
    expense_row = next((row for row in tb["rows"] if row["gl_account_id"] == expense_id), None)
    retained_row = next(row for row in tb["rows"] if row["gl_account_id"] == retained_earnings_id)
    assert revenue_row["net_balance"] == 0.0
    assert expense_row["net_balance"] == 0.0
    assert retained_row["net_balance"] == 600.0


def test_year_end_close_rejects_non_equity_retained_earnings_account(client, company):
    company_id = company["id"]
    revenue_id = _gl_account(client, company_id, "4501", "revenue")
    cash_id = _gl_account(client, company_id, "1501", "asset")
    _post(client, company_id, [{"gl_account_id": cash_id, "debit_amount": 100, "credit_amount": 0}, {"gl_account_id": revenue_id, "debit_amount": 0, "credit_amount": 100}], "2026-06-01")

    r = client.post(
        f"/year-end/close?company_id={company_id}",
        json={"start": "2026-01-01", "end": "2026-12-31", "retained_earnings_gl_account_id": cash_id},
    )
    assert r.status_code == 422


def test_year_end_close_with_nothing_to_close_is_rejected(client, company):
    company_id = company["id"]
    retained_earnings_id = _gl_account(client, company_id, "3502", "equity")
    r = client.post(
        f"/year-end/close?company_id={company_id}",
        json={"start": "2026-01-01", "end": "2026-12-31", "retained_earnings_gl_account_id": retained_earnings_id},
    )
    assert r.status_code == 422


def test_year_end_close_is_audited(client, company):
    company_id = company["id"]
    revenue_id = _gl_account(client, company_id, "4503", "revenue")
    cash_id = _gl_account(client, company_id, "1503", "asset")
    retained_earnings_id = _gl_account(client, company_id, "3503", "equity")
    _post(client, company_id, [{"gl_account_id": cash_id, "debit_amount": 100, "credit_amount": 0}, {"gl_account_id": revenue_id, "debit_amount": 0, "credit_amount": 100}], "2026-06-01")
    client.post(
        f"/year-end/close?company_id={company_id}",
        json={"start": "2026-01-01", "end": "2026-12-31", "retained_earnings_gl_account_id": retained_earnings_id},
    )
    entries = client.get(f"/audit-log?company_id={company_id}&entity_type=year_end_close").json()
    assert any(e["action"] == "close" for e in entries)


# --- Download all books -----------------------------------------------------


def test_books_export_returns_a_valid_workbook_with_all_sheets(client, company):
    company_id = company["id"]
    revenue_id = _gl_account(client, company_id, "4600", "revenue")
    cash_id = _gl_account(client, company_id, "1600", "asset")
    _post(client, company_id, [{"gl_account_id": cash_id, "debit_amount": 500, "credit_amount": 0}, {"gl_account_id": revenue_id, "debit_amount": 0, "credit_amount": 500}], "2026-01-10")

    r = client.get(f"/reports/books/export?company_id={company_id}&start=2026-01-01&end=2026-01-31")
    assert r.status_code == 200, r.text
    assert r.headers["content-type"].startswith("application/vnd.openxmlformats")

    workbook = load_workbook(io.BytesIO(r.content))
    assert set(workbook.sheetnames) == {"Journal", "Trial Balance", "Income Statement", "Balance Sheet"}
    journal_sheet = workbook["Journal"]
    assert journal_sheet.max_row >= 2  # header + at least one data row
